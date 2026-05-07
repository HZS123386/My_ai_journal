from io import BytesIO
from typing import Optional

# TXT 文件解析
def parse_txt(file_content: bytes) -> str:
    return file_content.decode('utf-8', errors='ignore')

# DOCX 文件解析
try:
    from docx import Document
    def parse_docx(file_content: bytes) -> str:
        doc = Document(BytesIO(file_content))
        return '\n'.join([paragraph.text for paragraph in doc.paragraphs])
except ImportError:
    def parse_docx(file_content: bytes) -> str:
        raise ImportError("需要安装 python-docx 库")

# XLSX 文件解析
try:
    from openpyxl import load_workbook
    def parse_xlsx(file_content: bytes) -> str:
        wb = load_workbook(BytesIO(file_content))
        text = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                row_text = '\t'.join([str(cell) if cell else '' for cell in row])
                if row_text.strip():
                    text.append(row_text)
        return '\n'.join(text)
except ImportError:
    def parse_xlsx(file_content: bytes) -> str:
        raise ImportError("需要安装 openpyxl 库")

# PDF 文件解析
try:
    from PyPDF2 import PdfReader
    def parse_pdf(file_content: bytes) -> str:
        reader = PdfReader(BytesIO(file_content))
        text = []
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text.append(page_text)
        return '\n'.join(text)
except ImportError:
    def parse_pdf(file_content: bytes) -> str:
        raise ImportError("需要安装 PyPDF2 库")

# 根据文件扩展名选择解析方法
def parse_file(file_content: bytes, filename: str) -> Optional[str]:
    if filename.lower().endswith('.txt'):
        return parse_txt(file_content)
    elif filename.lower().endswith('.docx'):
        return parse_docx(file_content)
    elif filename.lower().endswith('.xlsx'):
        return parse_xlsx(file_content)
    elif filename.lower().endswith('.pdf'):
        return parse_pdf(file_content)
    else:
        return None
        